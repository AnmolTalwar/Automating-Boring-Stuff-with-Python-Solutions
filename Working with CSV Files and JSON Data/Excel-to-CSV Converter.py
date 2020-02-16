import openpyxl, csv, os, re

os.chdir("C:\\Python\\Udemy\\Assignment\\Excel to CSV")

for excelFile in os.listdir('.'):

    if not excelFile.endswith('.xlsx'):
        continue
    
    wb = openpyxl.load_workbook(excelFile)

    for sheetName in wb.get_sheet_names():
         excelName = re.sub(".xlsx",'',excelFile)
         sheet = wb.get_sheet_by_name(sheetName)
         csvName = excelName+"_"+sheetName+'.csv'

         csvObj = open(csvName, 'w', newline='')
         csvWriter = csv.writer(csvObj)

         for rowNum in range(1, sheet.max_row + 1):
             rowData = []
             for colNum in range(1, sheet.max_column + 1):
                 rowData.append(sheet.cell(row=rowNum, column=colNum).value)

             csvWriter.writerow(rowData)

         csvObj.close()
        

print('Excel To CSV Done!!!!')
