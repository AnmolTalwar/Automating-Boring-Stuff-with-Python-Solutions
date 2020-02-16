import openpyxl
import os
from openpyxl.utils import get_column_letter
from pathlib import Path

path1 = "C:\\Python\\Udemy\\Assignment\\Text_Files"
os.chdir(path1)
p = Path(path1)

wb = openpyxl.Workbook()
sheet = wb.active

list=[]
for text in p.glob('*.txt'):
    x=open(text)
    list.append(x.readlines())


col=1
for file in list:
    row=1
    for sentence in file:
        sheet.cell(row=row, column=col).value = sentence
        row+=1
    
    col+=1

print("Text to Excel Done!!")
wb.save('Final_13_4.xlsx')




