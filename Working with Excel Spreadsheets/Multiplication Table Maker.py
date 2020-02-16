import sys,openpyxl,os

from openpyxl.utils import get_column_letter


n = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active

for i in range(2, n + 2):
    sheet['A' + str(i)] = i - 1
    
    colL = get_column_letter(i)
    sheet[colL + str(1)] = i - 1

for column in range(2, n + 2):
    for row in range(2, n + 2):
        colL1 = get_column_letter(column)
        sheet[colL1 + str(row)] = (column - 1) * (row - 1)

print("Sheet Created!!")
wb.save('mulExcercise1.xlsx')
