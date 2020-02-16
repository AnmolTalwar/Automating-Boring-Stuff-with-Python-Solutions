import sys
import os
import openpyxl

os.chdir('C:\\Python\\Udemy\\Assignment\\Excel to Text')

wb = openpyxl.load_workbook("Test.xlsx")
sheet = wb.active

#list=[[]]*sheet.max_column
list=[[] for i in range(sheet.max_column)]

#print(list)
for column in range(1, sheet.max_column+1):
    for row in range(1, sheet.max_row+1):
        #print(row)
        #print(column)
        list[column-1].append(sheet.cell(row=row, column=column).value)
        #print(list)


for i in range(len(list)):
    f=open('Column%s.txt'%str(i+1), 'w')
    a = ' '.join(list[i])
    f.write(a)
    f.close()

print('Copying Excel to Text Done!!')

