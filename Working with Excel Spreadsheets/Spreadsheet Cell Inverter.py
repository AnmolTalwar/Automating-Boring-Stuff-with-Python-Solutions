import sys
import os
import openpyxl

file = sys.argv[1]

os.chdir('C:\\Python\\Udemy\\Assignment')

wb = openpyxl.load_workbook(file)
sheet = wb.active

wb1 = openpyxl.Workbook()
sheet1 = wb1.active


for row in range(1, max(sheet.max_row,sheet.max_column)+ 1):
    for column in range(1, max(sheet.max_row,sheet.max_column) + 1):
            sheet1.cell(row=row, column=column).value = sheet.cell(row=column, column=row).value

wb1.save('Final_13_3.xlsx')
print('Transpose Done!!')
