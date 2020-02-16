import sys
import os
import openpyxl


start = int(sys.argv[1])
length = int(sys.argv[2])
file = sys.argv[3]

os.chdir('C:\\Python\\Udemy\\Assignment')

wb = openpyxl.load_workbook(file)
sheet = wb.active

wb1 = openpyxl.Workbook()
sheet1 = wb1.active


for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        if row < start:
            sheet1.cell(row=row, column=column).value = sheet.cell(row=row, column=column).value
        else:
            sheet1.cell(row=row+length, column=column).value = sheet.cell(row=row, column=column).value

wb1.save('Final_13_2.xlsx')

print('Blank Values Pasted!!')
